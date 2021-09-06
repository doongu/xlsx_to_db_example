from openpyxl import load_workbook, Workbook

def data_selecter(location):

    return_data = []

    for index_x in range(len(location)):
        for index_y in range(len(location[index_x])):
            # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
            load_wb = load_workbook(location[index_x][index_y], data_only=True)
            load_ws = load_wb.worksheets[0] # 첫 번째 워크시트 가져오기
            
            file_name = ""
            for i in range(len(location[index_x][index_y])-1, -1, -1):
                if location[index_x][index_y][i] != "/":
                    file_name += location[index_x][index_y][i]
                else:
                    break
            file_name = file_name[::-1]


            # "2021-09-02_심심수산_울도"를 키 값으로 넣어야할 듯
            return_list = []


            index = 2 # 2번째 행부터 데이터를 가져와야함
            while True:
                a = str(load_ws.cell(index, 1).value) # 쇼핑몰
                c = str(load_ws.cell(index, 3).value) # 주문일
                d = str(load_ws.cell(index, 4).value) # 온라인상품명
                e = str(load_ws.cell(index, 5).value) # 옵션
                f = str(load_ws.cell(index, 6).value) # 판매자관리코드
                g = str(load_ws.cell(index, 7).value) # 주문수량
                h = str(load_ws.cell(index, 8).value) # 금액
                i = str(load_ws.cell(index, 9).value) # 배송비
                j = str(load_ws.cell(index, 10).value) # 주문자명
                k = str(load_ws.cell(index, 11).value) # 수령자명
                l = str(load_ws.cell(index, 12).value) # 주소
                

                if a == "None" and c=="None" and d =="None" and e =="None" and f =="None" and g =="None" and h =="None" and i =="None" and j =="None" and k =="None" and l =="None":
                    break
                else:
                    return_data.append([file_name, a, c, d, e, f, g, h, i, j, k, l])
                    index+=1


    return return_data


# data_selecter("C:/Users/Hellllo/Desktop/크몽/xlsx_db연동/매출분석/2022/3월/NEW심심수산/2021-09-06_심심수산_울도.xlsx")