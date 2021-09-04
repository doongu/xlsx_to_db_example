from openpyxl import load_workbook, Workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/Users/Hellllo/Desktop/크몽/xlsx_db연동/매출분석/2022/3월/NEW심심수산/2021-09-02_심심수산_울도.xlsx", data_only=True)
load_ws = load_wb.worksheets[0] # 첫 번째 워크시트 가져오기
# print(str(load_ws.cell(2, 9).value))
# print(str(load_ws.cell(2, 1).value))
# print(str(load_ws.cell(2, 3).value))
# print(str(load_ws.cell(2, 4).value))
# print(str(load_ws.cell(2, 5).value))
# print(str(load_ws.cell(2, 6).value))
# print(str(load_ws.cell(2, 7).value))
# print(str(load_ws.cell(2, 8).value))
# print(str(load_ws.cell(2, 9).value)) # 대쉬가 들어가는 경우 마이너스로 인식해서 오류 ? 
# print(str(load_ws.cell(2, 10).value))
# print(str(load_ws.cell(2, 11).value))
# 셀 주소로 값 출력
for index in range(2, 3):
    a = str(load_ws.cell(index, 1).value)#상품명
    c = str(load_ws.cell(index, 3).value) #수량
    d = str(load_ws.cell(index, 4).value) #가격
    e = str(load_ws.cell(index, 5).value)
    f = str(load_ws.cell(index, 6).value)
    g = str(load_ws.cell(index, 7).value)
    h = str(load_ws.cell(index, 8).value)
    i = str(load_ws.cell(index, 9).value)
    j = str(load_ws.cell(index, 10).value)
    k = str(load_ws.cell(index, 11).value)
    l = str(load_ws.cell(index, 12).value)
    # #수량을 넣어서 check_name을 정한다.
    # check_name_alg(select_count)

    # #엑셀에서 공백이 있는데 이를 무시하고 값을 추가한다.
    # ignore_blank_alg(select_name, select_count, select_price)
print(a,c,d,e,f,g,h,i,j,k,l)